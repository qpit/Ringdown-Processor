from .Measurement import Measurement,AutoFitError,Property
from .MeasurementFileReader import InvalidRingdownFileError
import ntpath
from openpyxl import Workbook,load_workbook
import math
from . import serializer
from .constants import NMAX
from . import legacy_loader
from .Model_ExportFunctions import Model_ExportFunctions

class DuplicateMeasurementError(Exception):
    pass

class MeasurementMissingInfoError(Exception):
    def __init__(self,index,*args,**kwargs):
        self.index = index
        super().__init__(*args,**kwargs)

class Model(Model_ExportFunctions):
    """
    Main class for handling all processing of data in app.
    """

    @property
    def saved_measurements(self):
        return self._saved_measurements.copy()

    @property
    def num_of_measurements_to_be_processed(self) -> int:
        return len(self._measurements_to_be_processed)

    @property
    def measurements_to_be_processed(self) -> list:
        return self._measurements_to_be_processed.copy()

    @property
    def sampleIDs(self) -> list:
        l = list(self.sorted_to_sampleIDs.keys())
        l.sort()
        return l

    @property
    def modes(self) -> list:
        return self._uniquelist('mode')

    @property
    def design_types(self) -> list:
        l = list(self.sorted_to_design_types.keys())
        l.sort()
        return l

    @property
    def prefixes(self) -> list:
        l = list(self.sorted_to_prefixes.keys())
        l.sort()
        return l

    @property
    def properties(self) -> list:
        if not self._properties:
            self._get_properties()
        return self._properties

    @property
    def sorted_to_prefixes(self) -> dict:
        if not self._sorted_to_prefixes:
            self._get_prefixes()
        return self._sorted_to_prefixes

    @property
    def sorted_to_sampleIDs(self) -> dict:
        if not self._sorted_to_sampleIDs:
            self._get_sampleIDs()
        return self._sorted_to_sampleIDs

    @property
    def sorted_to_design_types(self) -> dict:
        if not self._sorted_to_design_types:
            self._get_design_types()
        return self._sorted_to_design_types

    def _uniquelist(self,attr:str):
        """
        Generates unique list on demand from list of saved measurements.
        :param attr:
        :return:
        """
        # _l = getattr(self, "_"+attr+"s")
        # if not _l and self._saved_measurements:
        #     # Find unique entries from list of saved measurements.
        #     _l = list(set([getattr(m,attr) for m in self._saved_measurements]))
        #     _l.sort()
        # _l_processed = getattr(self, "_"+attr+"s_processed")
        # if not _l_processed:
        #     return _l
        # l = _l + _l_processed
        # l.sort()
        # return l
        l = getattr(self, "_" + attr + "s")
        if not l and self._saved_measurements:
            # Find unique entries from list of saved measurements.
            l = list(set([getattr(m, attr) for m in self._saved_measurements+self._measurements_to_be_processed]))
            l.sort()
            # while '' in l:
            #     l.remove('')
        return l

    def get_properties_from_design_type(self,design_type:str) -> list:
        """
        Returns a typical list of properties for the given design_type
        :param design_type:
        :return:
        """
        if not design_type in self.sorted_to_design_types:
            return []

        # Obtain overview of existing parameters for design type
        d = {}
        measurements = self.sorted_to_design_types[design_type]
        for m in measurements:
            for p in m.properties:
                if not p.name in d:
                    d[p.name] = []
                d[p.name].append(p)

        # Keep parameters in use in the majority of measurements
        for name,l in d.items():
            if len(l) < len(measurements):
                del d[name]

        # For remaining parameters obtain the most common value
        common_properties = []
        for name, l in d.items():
            values = [p.value for p in l]
            common = max(set(values), key = values.count)
            common_properties.append(Property(name,common,l[0].unit))

        return common_properties

    def get_measurements_from_sampleID(self,sampleID:str) -> list:
        """
        Returns a list of measurements from a given sampleID
        :param sampleID:
        :return:
        """
        return [m for m in self._saved_measurements if m.sampleID == sampleID]

    def get_design_type_and_properties_from_sampleID(self,sampleID:str) -> tuple:
        """
        Returns a tuple of design type and a list of properties for a given sampleID if possible.
        :param sampleID:
        :return: (design_type, properties)
        """
        if sampleID in self.sampleIDs:
            # Sample ID already defined. Copy previous definitions.
            m = self.get_measurements_from_sampleID(sampleID)[0]
            l = []
            for p in m.properties:
                l.append(p.copy())
            return (m.design_type,l)

        else:
            prefix = self.name2prefix(sampleID)
            if prefix in self.sorted_to_prefixes:
                # Prefix already defined. Copy typical definitions.
                design_type = self.sorted_to_prefixes[prefix][0].design_type
                properties = self.get_properties_from_design_type(design_type)
                return (design_type,properties)

            # Return nothing
            return ('',[])

    def __init__(self):
        super().__init__()
        self._loaded_ringdown_files = [] # List of paths for loaded ringdown-containing-files.
        self._loaded_legend_files = [] # List of paths for loaded legend-containing-files.
        self._measurements_to_be_processed = [] # List of measurements being processed.
        self._saved_measurements = [] # List of measurements which have been saved. This list can be huge.
        self._lastID = 0 # ID of last saved measurement.

        ''' Unique lists of already defined sampleIDs, modes, design types 
        and parameters used to aid during dataprocessing. '''
        # From list of saved measurements.
        self._modes = []
        self._properties = []

        # Dictionary of measurements sorted into prefixes, sample IDs and design types as the key.
        self._sorted_to_prefixes = {}
        self._sorted_to_sampleIDs = {}
        self._sorted_to_design_types = {}

        # # From the current list of measurements to be processed.
        # self._sampleIDs_processed = []
        # self._design_types_processed = []
        # self._modes_processed = []
        # self._parameters_processed = []

    def _get_properties(self):
        """
        Updates the _parameters variable to contain a list of all used properties and their most common values.
        :return:
        """
        d = {}
        for m in self._saved_measurements:
            if m.properties:
                for p in m.properties:
                    if not p.name in d:
                        d[p.name] = []
                    d[p.name].append(p)
        properties = []
        for name,l in d.items():
            l_values = [p.value for p in l]
            commonvalue = max(set(l_values), key = l_values.count)
            l_units = [p.unit for p in l]
            commonunit = max(set(l_units), key=l_units.count)
            properties.append(Property(name=name,value=commonvalue,unit=commonunit))

        self._properties = properties

    def _get_prefixes(self):
        """
        Returns a list of prefixes and their design type assuming the format LLLNN-NN, where L and N are letters and numbers, respectively.
        :return:
        """
        d = {}
        for m in self._saved_measurements:
            prefix = self.name2prefix(m.sampleID)
            if prefix:
                if not prefix in d:
                    d[prefix] = []
                d[prefix].append(m)
        self._sorted_to_prefixes = d

    def _get_sampleIDs(self):
        """
        Returns a list of prefixes and their design type assuming the format LLLNN-NN, where L and N are letters and numbers, respectively.
        :return:
        """
        d = {}
        for m in self._saved_measurements:
            if m.sampleID:
                if not m.sampleID in d:
                    d[m.sampleID] = []
                d[m.sampleID].append(m)
        self._sorted_to_sampleIDs = d

    def _get_design_types(self):
        """
        Returns a list of prefixes and their design type assuming the format LLLNN-NN, where L and N are letters and numbers, respectively.
        :return:
        """
        d = {}
        for m in self._saved_measurements:
            if m.design_type:
                if not m.design_type in d:
                    d[m.design_type] = []
                d[m.design_type].append(m)
        self._sorted_to_design_types = d

    @staticmethod
    def name2prefix(name:str) -> str:
        i = 0
        while i < len(name) - 1:
            if name[i].isnumeric():
                break
            i += 1
        if i > 0 and i < len(name) - 1:
            # Valid prefix
            return name[0:i]
        return ''

    def load(self):
        """
        Loads relevant files if they exist.
        :return:
        """
        serializer.deserialize_model(self)

    def save(self):
        """
        Loads relevant files if they exist.
        :return:
        """
        serializer.serialize_model(self)


    def load_measurement_file(self, path:str):
        """
        Loads the file for dataprocessing.
        :param filename:
        :return:
        """
        filename = ntpath.basename(path)
        extension = filename.split('.')[-1].lower()
        if extension in ['xls', 'xlsx']:
            # Possibly legend file.
            self._loaded_legend_files.append(path)
        else:
            # Possibly a ringdown file.
            self._loaded_ringdown_files.append(path)

    def process_loaded_files(self):
        """
        Processes all loaded files and clears the loaded files lists. Adds to list of processing measurements
        :return:
        """

        ''' Process all legend files '''
        legend_entries = []
        for path in self._loaded_legend_files:
            legend_entries.extend(self._process_legend_file(path))

        ''' Process all ringdown files '''
        errors = []
        for path in self._loaded_ringdown_files:
            filename = ntpath.basename(path)
            measurement = Measurement()

            # Find matching legend entry, if it exists.
            filename_no_ext = ''.join(filename.split('.')[0:-1])
            for legend_entry in legend_entries:
                if legend_entry['filename'].lower() == filename_no_ext.lower():
                    # Matching entry found.
                    measurement.sampleID = legend_entry['sample']
                    measurement.mode = legend_entry['mode']
                    measurement.pressure = legend_entry['pressure']
                    measurement.notes = legend_entry['notes'].strip()
                    measurement.temperature = legend_entry['temperature']
                    measurement.is_bad = legend_entry['isbad']

            # Process file
            is_ok = False
            try:
                measurement.loadfile(path)

                # Check if duplicate already exists and saved.
                if self.check_duplicate_measurement(measurement):
                    raise DuplicateMeasurementError("Duplicate of measurement file '{:s}' already exists.".format(path))

                self._measurements_to_be_processed.append(measurement)
                is_ok = True
            except (InvalidRingdownFileError, DuplicateMeasurementError) as e:
                errors.append(e)

            if is_ok:
                try:
                    measurement.autofit()
                except AutoFitError as e:
                    errors.append(e)

        # Reset list of loaded files.
        self._loaded_legend_files.clear()
        self._loaded_ringdown_files.clear()

        # Infer missing info where possible
        for m in self._measurements_to_be_processed:
            self.fillout_measurement(m)

        # Return overview of errors
        return errors

    def fillout_measurement(self,measurement:Measurement):
        """
        Fills out missing information where possible. Returns True if measurement has been updated.
        :param measurement:
        :return:
        """
        if measurement.sampleID:
            design_type,properties = self.get_design_type_and_properties_from_sampleID(measurement.sampleID)
            if design_type:
                measurement.design_type = design_type
                measurement.clear_properties()
                for p in properties:
                    measurement.add_property(p)
                return True
        return False

    def check_duplicate_measurement(self,measurement:Measurement) -> bool:
        """
        Returns True if the given measurement already exists in the saved list of measurements.
        It checks datetime.
        :param measurement:
        :return:
        """
        for m in self._saved_measurements:
            if m.datetime == measurement.datetime:
                return True
        return False

    def clear_measurements_to_be_processed(self):
        """
        Clear the list of measurements to be processed.
        :return:
        """
        self._loaded_legend_files.clear()
        self._loaded_ringdown_files.clear()
        self._measurements_to_be_processed.clear()

    @staticmethod
    def _process_legend_file(path:str):
        """
        Processes legend file and returns a list of legend data
        :param path:
        :return:
        """

        l = []

        wb = load_workbook(filename=path)
        sheet = wb.active

        # Find start of column names
        row0 = 1
        found = False
        for row in range(1,1+NMAX):
            cval = sheet.cell(row=row,column=1).value
            if not cval is None:
                if "sample" in cval.lower():
                    row0 = row
                    found = True
                    break

        if not found:
            raise Exception('Not a valid legend file')

        ''' Scan column names for specific entries '''
        # Find width of table
        col = 1
        max_column = 1
        while col < NMAX:
            if sheet.cell(row=row0,column=col).value is None:
                max_column = col-1
                break
            col += 1

        # Find number of rows
        row = row0+1
        while row < NMAX:
            found = False
            for col in range(1,max_column+1):
                if sheet.cell(row=row,column=col).value is not None:
                    found = True
                    break
            if not found:
                break
            row += 1

        max_row = row-1

        # Initialize dictionary containing column indexes
        coldata = dict()
        atts = ['sample', 'filename', 'mode', 'pressure', 'notes', 'temperature']
        for att in atts:
            # if att in ['pressure']:
            #     coldata[att] = math.nan
            # if att in ['temperature']:
            #     coldata[att] = 293 # [K] Most likely temperature.
            # if att in ['notes']:
            #     coldata[att] = ''
            # else:
            #     coldata[att] = None
            coldata[att] = None

        # Identify position of columns.
        for col in range(1,1+max_column):
            cval = sheet.cell(row=row0, column=col).value.lower()
            for att in atts:
                if att in cval:
                    coldata[att] = col

        # Critical information: Need at least filename.
        if coldata['filename'] == None:
            raise Exception('Legend file contains no filenames.')

        ''' Scan rows until none is left '''
        # Initialize row data.
        rowdata = dict()
        for att in atts:
            rowdata[att] = None
        rowdata['temperature'] = 293 # [K] Default value for temperature, as this is the usual case.
        rowdata['mode'] = ''
        rowdata['empty_row'] = True
        rowdata['skip'] = False
        rowdata['isbad'] = False

        # Row loop
        for row in range(row0 + 1, 1+max_row):
            rowdata['empty_row'] = True
            rowdata['skip'] = False

            # Method for reading
            def read(key: str, skipblanck=True):
                col = coldata[key]
                if col == None:
                    return
                cval = sheet.cell(row=row, column=col).value
                # if sheet.cell_type(row, col) == 2:
                #     # Number. Convert to string properly.
                #     if cval.is_integer():
                #         cval = str(int(cval))
                #     else:
                #         cval = str(cval)
                # else:
                #     cval = str(cval)

                if cval is None:
                    if not skipblanck:
                        rowdata[key] = ""
                elif cval == "-":
                    rowdata["empty_row"] = False
                    rowdata[key] = ""
                else:
                    rowdata["empty_row"] = False
                    rowdata[key] = cval

                    if key == 'sample':
                        # Reset bad sample mark.
                        rowdata['isbad'] = False

                    # Check if marked as bad sample
                    if key == 'notes':
                        if 'bad' in cval.lower():
                            # Most likely marked as bad sample.
                            rowdata['isbad'] = True

            # Read data. Use previous read data if empty.
            if sheet.cell(row=row, column=coldata['filename']).value == "-":
                rowdata['skip'] = True
            for att in atts:
                if att in ["notes"]:
                    read(att, skipblanck=False)
                else:
                    read(att)

            if rowdata['skip']:
                continue
            elif rowdata['empty_row']:
                break

            # Save rowdata to list
            def str2float(s: str) -> float:
                try:
                    return float(s)
                except:
                    return math.nan

            if rowdata['notes'] is None:
                rowdata['notes'] = ''

            l.append(
                {'filename': rowdata['filename'],
                 'sample': rowdata['sample'],
                 'notes': rowdata['notes'],
                 'pressure': str2float(rowdata['pressure']),
                 'mode': str(rowdata['mode']),
                 'temperature': str2float(rowdata['temperature']),
                 'isbad': rowdata['isbad']
                 }
            )

        # Return list of legend entries.
        return l

    def remove_loaded_measurement(self,measurement:Measurement):
        """
        Removes the given measurement from dataprocessing
        :param measurement:
        :return:
        """
        self._measurements_to_be_processed.remove(measurement)

    def save_processed_measurements(self):
        """
        Saves all measurements which were processed.
        :return:
        """
        errors = []
        for i in reversed(range(len(self._measurements_to_be_processed))):
            m = self._measurements_to_be_processed[i]
            if m.ready_for_save:
                self.save_measurement(m)
                self._measurements_to_be_processed.remove(m)
            else:
                errors.append(MeasurementMissingInfoError(i,"Measurement file '{:s}' is missing a valid sample ID.".format(m.path)))

        self._changed()
        self.save()

        return errors

    def save_measurement(self, m:Measurement):
        """
        Saves the given measurement.
        :param measurement:
        :return:
        """
        while self._measurement_ID_exists(self._lastID):
            self._lastID += 1
        m.set_measurement_ID(self._lastID)
        self._saved_measurements.append(m)
        m.save2datafile()

    def _measurement_ID_exists(self,ID:int):
        for m in self._saved_measurements:
            if m.measurement_ID == ID:
                return True
        return False

    # def _save2files(self):
    #     """
    #     Saves data to files.
    #     :return:
    #     """
    #     for m in self._saved_measurements:
    #         if m.changed:
    #             m.save2datafile()

    def _changed(self):
        """
        Is called whenever saved data is changed. Resets parts of the object.
        :return:
        """
        self._modes.clear()
        self._properties.clear()
        self._sorted_to_prefixes.clear()
        self._sorted_to_sampleIDs.clear()
        self._sorted_to_design_types.clear()

    def load_legacy_data(self,path):
        """
        Loads old data based on the older software.
        :param path:
        :return:
        """
        legacy_loader.load(self,path)
