"""
An addin class for the Model class focused on export functionality
"""
from openpyxl import Workbook
from .Measurement import ATTRIBUTES_BASIC_INFORMATION
from .measurement_filters import filt_design_type,filt_sampleID,filt_property

class Model_ExportFunctions():

    def export_csv(self,
               export_path=None,
               design_type = None,
               sampleID_substring = None,
               include_bad = False,
               include_gas_limited = False,
               sort_designs_into_sheets = True,
               ):
        """
        Exports data to csv file.
        :param design_type: Choose design type. If none, all design types are used.
        :param sampleID_substring: Filter sampleID by substring.
        :param include_bad: If True, includes bad data.
        :param include_gas_limited: If True, includes gas-limited data.
        :param export_path: If none, saves as 'export.xlsx' at current folder.
        :param sort_designs_into_sheets: If true, defines a sheet for each selected design type.
        :return:
        """

        ''' Filter data '''
        filters = []

        # Setup filters
        if design_type:
            filters.append(filt_design_type(design_type))
        if sampleID_substring:
            filters.append(filt_sampleID(sampleID_substring))
        if not include_bad:
            filters.append(filt_property('is_bad',False,'=='))
        if not include_gas_limited:
            filters.append(filt_property('is_gas_limited', False, '=='))

        # Perform filtering
        measurements = [m for m in self.saved_measurements if all(f(m) for f in filters)]

        ''' Create workbook and its sheets '''
        wb = Workbook()  # Create workbook
        ws = wb.active  # Grab the active worksheet

        def print_sheet(measurements,ws):
            # Obtain unique list of parameters from filtered list
            property_names = []
            for m in measurements:
                for p in m.properties:
                    t = (p.name, p.unit)
                    if not t in property_names:
                        property_names.append(t)
            property_names.sort()

            # Write columns
            column = 1
            row = 1

            basic_info = []
            for attr,label in ATTRIBUTES_BASIC_INFORMATION:
                # if label in ['notes']:
                #     continue
                basic_info.append((attr,label))
                if attr == 'quality_factor':
                    basic_info.append(('qf', 'Qf [Hz]'))

            for attr, label in basic_info:
                ws.cell(row=row, column=column, value=label)
                column += 1

            column_properties = column
            for name, unit in property_names:
                value = name
                if unit:
                    value += ' [' + unit + ']'
                ws.cell(row=row, column=column, value=value)
                column += 1

            # Write data
            for m in measurements:
                row += 1
                column = 1

                # Basic info
                for attr, label in basic_info:
                    if attr == 'qf':
                        value = m.quality_factor * m.frequency
                    else:
                        value = getattr(m, attr)
                    ws.cell(row=row, column=column, value=value)
                    column += 1

                # Properties
                for p in m.properties:
                    t = (p.name, p.unit)
                    i = property_names.index(t)
                    ws.cell(row=row, column=column_properties + i, value=m[p.name])

        if sort_designs_into_sheets:
            design_types = [m.design_type for m in measurements]
            design_types = list(set(design_types))
            for i_design,design_type in enumerate(design_types):
                if design_type == '':
                    design_type = 'noname'
                if i_design == 0:
                    ws.title = design_type
                else:
                    wb.create_sheet(design_type)
                ws = wb[design_type]
                print_sheet([m for m in measurements if m.design_type == design_type],
                            ws)
        else:
            print_sheet(measurements,ws)

        ''' Save to file '''
        if export_path is None:
            export_path = 'export.xlsx'
        wb.save(export_path)